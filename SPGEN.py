### SPREAD SHEET GENERATOR ###

from openpyxl import Workbook
from typing import List
from UNGEN import list_usernames
from PARSR import list_column, stringify
import random
import string

## Configs ##
pwlen = 8

def create_pw(n: int) -> str:
  pw = ''.join(random.choice(string.ascii_letters + string.digits) for i in range(n))
  return pw

def merge_lists(m: List[str], n: List[str], p: List[str]) -> List[List[str]]:
  length = max(len(l) for l in [m,n,p])
  for l in [m,n,p]:
    if len(l) < length:
      for i in range(length - len(l)):
       l.append('')
  merged = [ [ m[i], n[i], p[i] ] for i in range(length) ]
  return merged


def rename_users(rpath: str, wpath: str, col: list) -> None:
  ## Worksheet Configs ##
  wb = Workbook()
  mill = wb.create_sheet("millenium users", 0)
  lost = wb.create_sheet("users not in ad", 1)

  #full_names, c_names, c_idnum = [
  #    list_column(rpath,col[i]) for i in range(len(col))]

  full_names  = list_column(rpath, col[0])
  c_names     = stringify(list_column(rpath, col[1]))
  c_idnums    = stringify(list_column(rpath, col[2]))
  # list_usernames takes a file path and a column to parse
  # returns a list of millenium usernames that match existing AD SAM account names
  # and a list of usernames that do not exist in AD, but do in millenium
  mill_users, lost_users = list_usernames(full_names,col[0])
  mill.append(['Cerner ID','Cerner Username','AD Username','Password'])
  lost.append(['Username'])
  cred = []

  merged = merge_lists(c_idnums, c_names, mill_users)

  #def byname(l: List[str]):
  #  return l[2]

  #merged = merged.sort(key=byname)

  for i in merged:
    mill.append(i)

  for uname in lost_users:
    uname = [uname]
    lost.append(uname)

  wb.save(wpath)
