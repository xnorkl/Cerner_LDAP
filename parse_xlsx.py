## Parse XLSX ##
## TODO ##
# FIRST: Make REGEX to Parse Names into a list of 3-tuple NAMES(First,Last,Middle)
# THEN: Outline the functions needed to parse the data from the given xlsx files
#       so that the final output is a dictionary USERNAMES{'IN_AD','NOT_IN_AD'} 
from bisect import insort
from openpyxl import load_workbook
from typing import List,Dict
import re
import pprint 

# Set PPrint
pp = pprint.PrettyPrinter(indent=1)

# Set workbook (path can be improved)
path = 'Providers.xlsx'
wb = load_workbook(path)
ws = wb.active
 
# Full Name column of Workbook
full_name = ws['B']

def sanitize(names: List[str]) -> List[str]:
  ## Sanitize list ##
  # Tokens to remove (easier than using RE to remove "()"
  de_s  = "(),_."
  clean = str.maketrans(dict.fromkeys(de_s))

  # Replace all unwanted strings with a blank space 
  sanitized: List[str] = sorted([re.sub(
    r'(Test\s.*)|(Cerner)|(Physician)|(Advisor)|(\sHw)|(De La)|(^St\s)|[A-Z]{2,}|\b[A-Z]?-\S+|(\s)(?=\s)',
    # Remove any Names containing Test
    # Remove any (<WORD>)
    # [A-Z]{2,} : Remove any title (MD,PHD,etc)
    # \b[A-Z]?-\S+ : Remove '-' and any following word
    # Remove the last word if not an initial
    '',
    str(n.value).translate(clean)).strip()
    for n in names
    ])
  return sanitized

def record(names: List[str]) -> List[str]:#Dict[int,list]:
  rubrick = {"First","Last","M"}
  split_names = list(
      sub.append(' ') or sub if len(sub) == 2 else sub 
      for sub in list(
        re.split(r'\W+', name) 
        for name in names[1:] if name)
      )

  return split_names

names = record(sanitize(full_name))
pp.pprint(names)
#rubrick = ['First','Last','Mid']
# Split each name into sublists
#split_names = dict(re.split(r'\W+', name) for name in full_names[1:] if name)
#pp.pprint(split_names)

#first = [sublist[0] for sublist in split_names]
#last  = [sublist[1] for sublist in split_names]
#mid   = [sublist[-1] for sublist in split_names]

#mid = [name for name in mid if name]

#for n in mid:
#  if n in last:
#    mid.remove(n)
  

#Name_Dict = dict(zip(['First','Last','Mid'],[first,last,mid]))
#pp.pprint(Name_Dict)

#New_Name = list(zip(first,last,mid))

#pp.pprint(New_Name)

#initial   = re.compile(r'\b[A-A]')
#tail      = re.compile(r'\b[a-zA-Z]+')
#f_initial = [str(initial.findall(name)).strip for name in first]
#f_tail     = [tail.findall(name) for name in first]

#usernames = ["{}{}".format(i,j) for i,j in zip(last, f_initial)]

#for u in usernames:
#  print(u)
